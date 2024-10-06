import openpyxl
from datetime import datetime
from typing import List
from cafe_management.src.models import Category, Customer, Invoice, Order, OrderItem, Product
# from cafe_management.src.models.staff import Staff

class DataController:
    @staticmethod
    def load_categories(file_path: str) -> List[Category]:
        categories = []
        if file_path.endswith(".xlsx"):
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                category_id, name = row
                if category_id and name:
                    categories.append(Category(name=name, id=category_id))
        return categories

    @staticmethod
    def save_categories(file_path: str, categories: List[Category]):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Id", "Name"])
        for category in categories:
            sheet.append([category.id, category.name])
        wb.save(file_path)

    @staticmethod
    def load_products(file_path: str) -> List[Product]:
        products = []
        if file_path.endswith(".xlsx"):
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                product_id, name, category_id, price = row
                if product_id and name and category_id and price:
                    products.append(
                        Product(
                            name=name,
                            category_id=category_id,
                            price=price,
                            id=product_id,
                        )
                    )
        return products

    @staticmethod
    def save_products(file_path: str, products: List[Product]):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Id", "Name", "CategoryId", "Price"])
        for product in products:
            sheet.append([product.id, product.name, product.category_id, product.price])
        wb.save(file_path)

    @staticmethod
    def load_customers(file_path: str) -> List[Customer]:
        customers = []
        if file_path.endswith(".xlsx"):
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                customer_id, name, birthday_str, phone_number, email, points = row
                if customer_id and name and birthday_str:
                    try:
                        birthday = datetime.strptime(birthday_str, "%Y-%m-%d")
                    except ValueError:
                        birthday = datetime.min
                    customers.append(
                        Customer(
                            name=name,
                            birthday=birthday,
                            phone_number=phone_number,
                            email=email,
                            id=customer_id,
                            points=points,
                        )
                    )
        return customers

    @staticmethod
    def save_customers(file_path: str, customers: List[Customer]):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Id", "Name", "Birthday", "PhoneNumber", "Email", "Points"])
        for customer in customers:
            sheet.append(
                [
                    customer.id,
                    customer.name,
                    customer.birthday.strftime("%Y-%m-%d"),
                    customer.phone_number,
                    customer.email,
                    customer.points,
                ]
            )
        wb.save(file_path)

    @staticmethod
    def load_orders(file_path: str) -> List[Order]:
        orders = []
        if file_path.endswith(".xlsx"):
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                order_id, customer_id, order_date_str, items_str, points = row
                if order_id and customer_id and order_date_str and items_str:
                    try:
                        order_date = datetime.strptime(order_date_str, "%Y-%m-%d")
                    except ValueError:
                        order_date = datetime.min
                    items = []
                    item_parts = items_str.split(",")
                    for item_part in item_parts:
                        product_id, quantity, unit_price = map(
                            float, item_part.split(":")
                        )
                        items.append(
                            OrderItem(
                                product_id=product_id,
                                quantity=quantity,
                                unit_price=unit_price,
                            )
                        )
                    orders.append(
                        Order(
                            id=order_id,
                            customer_id=customer_id,
                            order_date=order_date,
                            items=items,
                            points=points,
                        )
                    )
        return orders

    @staticmethod
    def save_orders(file_path: str, orders: List[Order]):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Id", "CustomerId", "OrderDate", "Items", "Points"])
        for order in orders:
            items_str = ",".join(
                [
                    f"{item.product_id}:{item.quantity}:{item.unit_price}"
                    for item in order.items
                ]
            )
            sheet.append(
                [
                    order.id,
                    order.customer_id,
                    order.order_date.strftime("%Y-%m-%d"),
                    items_str,
                    order.points,
                ]
            )
        wb.save(file_path)

    @staticmethod
    def load_invoices(file_path: str) -> List[Invoice]:
        invoices = []
        if file_path.endswith(".xlsx"):
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                invoice_id, order_id, date_str = row
                if invoice_id and order_id and date_str:
                    try:
                        date = datetime.strptime(date_str, "%Y-%m-%d")
                    except ValueError:
                        date = datetime.min
                    invoices.append(
                        Invoice(id=invoice_id, order_id=order_id, date=date)
                    )
        return invoices

    @staticmethod
    def save_invoices(file_path: str, invoices: List[Invoice]):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Id", "OrderId", "Date"])
        for invoice in invoices:
            sheet.append(
                [invoice.id, invoice.order_id, invoice.date.strftime("%Y-%m-%d")]
            )
        wb.save(file_path)